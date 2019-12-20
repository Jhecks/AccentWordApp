import sys
import subprocess
import time
from PyQt5 import QtWidgets, QtCore
import AccentWord
import Inserter
import pymysql
from pymysql.cursors import DictCursor
import random
from openpyxl import load_workbook, Workbook

dictionary = [0]*2
dictionary[0] = []
dictionary[1] = []


def backup():
    global intDB
    try:
        intDB = load_workbook('DB.xlsx')
    except Exception as inst:
        intDB = Workbook()
        intDB.save("DB.xlsx")
        intDB = load_workbook('DB.xlsx')
    global sheet
    sheet = intDB.active
    for column in "BC":
        for row in range(2, len(dictionary[0])):
            cell_name = "{}{}".format(column, row)
            if column == "B":
                sheet[cell_name].value = dictionary[0][row-2]
            else:
                sheet[cell_name].value = dictionary[1][row-2]
    intDB.save("DB.xlsx")


def AltInit():
    global intDB
    intDB = load_workbook('DB.xlsx')
    global sheet
    sheet = intDB.active

    for column in "BC":
        for row in range(2, sheet.max_row + 1):
            cell_name = "{}{}".format(column, row)
            if column == "B":
                dictionary[0].append(sheet[cell_name].value)
            else:
                dictionary[1].append(sheet[cell_name].value)


def XlsxInsert(InputText):
    sheet.append((0, InputText.lower(), InputText))
    intDB.save("DB.xlsx")
    dictionary[0].append(InputText.lower())
    dictionary[1].append(InputText)


def initialization():
    global db
    try:
        db = pymysql.connect(
            host='remotemysql.com',
            user='slKnUCkyMM',
            password='kxG08Jh68n',
            db='slKnUCkyMM',
            charset='utf8mb4',
            cursorclass=DictCursor)
        print('first')
        with db.cursor() as cursor:
            query = """UPDATE tDictionary SET date = %s"""
            date = time.strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute(query, date)
            db.commit()
    except Exception as inst:
        db = 0
        print(type(inst))
        print(inst.args)
        print(inst)
        if db == 0:
            AltInit()

    if db != 0:
        with db.cursor() as cursor:
            query = """
            SELECT
                word
            FROM
                tDictionary
            """
            if cursor.execute(query):
                for row in cursor:
                    dictionary[0].append(row['word'])

        with db.cursor() as cursor:
            query = """
            SELECT
                word_with_accent
            FROM
                tDictionary
            """
            if cursor.execute(query):
                for row in cursor:
                    dictionary[1].append(row['word_with_accent'])

        backup()


def Finder(InputText):
    try:
        if InputText == InputText.lower():
            for i in range(len(dictionary[0])):
                if dictionary[0][i] == InputText:
                    return 1, i, 'first'
        else:
            for i in range(len(dictionary[0])):
                if dictionary[1][i] == InputText or dictionary[1][i] == InputText + '*':
                    return 1, i, 'second'
                elif dictionary[0][i] == InputText.lower():
                    return 1, i, 'third'
        return 0, 0
    except Exception as inst:
        print(type(inst))
        print(inst.args)
        print(inst)


def DBInsert(InputText):
    with db.cursor() as cursor:
        query = """INSERT INTO tDictionary (word, word_with_accent) VALUES (lower(%s), %s)"""
        value = (InputText, InputText)
        cursor.execute(query, value)
        db.commit()
        dictionary[0].append(InputText.lower())
        dictionary[1].append(InputText)


def get_data(model):
    model.setStringList(dictionary[0])


class Inserter(QtWidgets.QMainWindow, Inserter.Ui_Dialog):
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self)
        self.setupUi(self)
        self.lineEdit.setPlaceholderText("Введите слово")
        self.pushButton.clicked.connect(self.Button1)
        self.lineEdit.returnPressed.connect(self.Button1)

    def Button1(self):
        try:
            InputText = self.lineEdit.text()
            if InputText == '':
                self.lineEdit.setPlaceholderText("Введите слово")
                return
            if db != 0:
                DBInsert(InputText)
            else:
                XlsxInsert(InputText)
            self.lineEdit.clear()
            model.setStringList(dictionary[0])
        except Exception as inst:
            print(type(inst))
            print(inst.args)
            print(inst)
            if inst.args[0] == 1062:
                self.lineEdit.setText("Слово уже существует")
            else:
                XlsxInsert(InputText)


class App(QtWidgets.QMainWindow, AccentWord.Ui_MainWindow):
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self)
        self.setupUi(self)
        self.lineEdit.setPlaceholderText("Введите слово")
        completer = QtWidgets.QCompleter()
        self.lineEdit.setCompleter(completer)
        global model
        model = QtCore.QStringListModel()
        completer.setModel(model)
        get_data(model)
        self.pushButton.clicked.connect(self.Button1)
        self.pushButton_2.clicked.connect(self.Button2)
        self.lineEdit.returnPressed.connect(self.Button1)
        self.action.triggered.connect(self.Insert)
        self.action_2.triggered.connect(self.Info)

    def Button1(self):
        InputText = self.lineEdit.text()
        if InputText == '':
            self.textBrowser.setText('Введите слово')
            return
        result = Finder(InputText)
        if result[0] == 1:
            if result[2] == 'first':
                self.textBrowser.setText(dictionary[1][result[1]])
            elif result[2] == 'second':
                output = 'Правильно (' + dictionary[1][result[1]] + ')'
                self.textBrowser.setText(output)
            else:
                output = 'Не верно (' + dictionary[1][result[1]] + ')'
                self.textBrowser.setText(output)
        else:
            self.textBrowser.setText('Слово не найдено')

    def Button2(self):
        random_num = int(random.uniform(0, len(dictionary[0])))
        self.lineEdit.setText(dictionary[0][random_num])
        self.textBrowser.setText(dictionary[1][random_num])

    def Insert(self):
        self.InserterWindow = Inserter()
        self.InserterWindow.show()

    def Info(self):
        subprocess.Popen(['notepad.exe', "help.txt"])


def main():
    try:
        app = 0
        initialization()
        app = QtWidgets.QApplication(sys.argv)
        window = App()
        window.show()
        app.exec_()
        if db != 0:
            db.close()
    except Exception as inst:
        print(type(inst))
        print(inst.args)
        print(inst)


if __name__ == '__main__':
    main()
